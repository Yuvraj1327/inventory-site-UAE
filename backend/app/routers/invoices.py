import io
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import Response
from pydantic import BaseModel
from typing import List, Optional

from app.core.security import require_staff_or_admin
from app.core.supabase_client import get_service_client
from app.services.compat import with_legacy_id, clean_list, now_iso, find_party_id_by_name

router = APIRouter(prefix="/api/invoices", tags=["invoices"])


class InvoiceItem(BaseModel):
    product_id: Optional[str] = None
    name: str = ""
    sku: str = ""
    qty: float = 0.0
    unit_price: float = 0.0


class InvoiceCreate(BaseModel):
    invoice_number: str
    customer: str = ""
    date: Optional[str] = None
    items: List[InvoiceItem] = []
    tax_percent: float = 0.0
    status: str = "unpaid"
    notes: str = ""
    ship_type: str = "EX-STOCK"
    currency: str = "AED"


def _totals(items: List[InvoiceItem], tax_percent: float):
    subtotal = sum(i.qty * i.unit_price for i in items)
    tax = subtotal * (tax_percent / 100.0)
    return round(subtotal, 2), round(tax, 2), round(subtotal + tax, 2)


def _to_legacy(invoice: dict, lines: list[dict]) -> dict:
    row = with_legacy_id(invoice)
    row["date"] = invoice.get("invoice_date")
    row["items"] = [
        {"product_id": li.get("product_id"), "name": li.get("description", ""),
         "sku": li.get("part_number", ""), "qty": li.get("qty", 0), "unit_price": li.get("unit_price", 0)}
        for li in lines
    ]
    return row


async def _available_qty(sb, product_id: str) -> float:
    r = sb.table("inventory").select("available_qty").eq("product_id", product_id).execute()
    return (r.data[0]["available_qty"] if r.data else 0) or 0


async def _bump_inventory(sb, product_id: str, delta: float, invoice_id: str):
    current = await _available_qty(sb, product_id)
    new_qty = current + delta
    existing = sb.table("inventory").select("product_id").eq("product_id", product_id).execute().data
    if existing:
        sb.table("inventory").update({"available_qty": new_qty, "updated_at": now_iso()}).eq("product_id", product_id).execute()
    else:
        sb.table("inventory").insert({"product_id": product_id, "available_qty": max(new_qty, 0)}).execute()
    sb.table("inventory_movements").insert({
        "product_id": product_id, "movement_type": "sale" if delta < 0 else "return",
        "quantity": delta, "reference_type": "invoice", "reference_id": invoice_id,
    }).execute()


@router.get("")
async def list_invoices(staff=Depends(require_staff_or_admin)):
    sb = get_service_client()
    invoices = sb.table("invoices").select("*").order("invoice_date", desc=True).execute().data or []
    if not invoices:
        return []
    ids = [i["id"] for i in invoices]
    lines = sb.table("invoice_lines").select("*").in_("invoice_id", ids).execute().data or []
    by_id = {}
    for li in lines:
        by_id.setdefault(li["invoice_id"], []).append(li)
    return [_to_legacy(i, by_id.get(i["id"], [])) for i in invoices]


@router.post("")
async def create_invoice(payload: InvoiceCreate, staff=Depends(require_staff_or_admin)):
    sb = get_service_client()
    dup = sb.table("invoices").select("id").eq("invoice_number", payload.invoice_number).execute()
    if dup.data:
        raise HTTPException(status_code=400, detail="Invoice number already exists")

    for it in payload.items:
        if it.product_id:
            avail = await _available_qty(sb, it.product_id)
            if avail < it.qty:
                raise HTTPException(status_code=400, detail=f"Not enough stock for {it.name}: {avail} available")

    subtotal, tax, total = _totals(payload.items, payload.tax_percent)
    customer_id = await find_party_id_by_name(sb, "customers", payload.customer)
    header = {
        "invoice_number": payload.invoice_number, "customer": payload.customer, "customer_id": customer_id,
        "invoice_date": (payload.date or now_iso())[:10], "subtotal": subtotal, "tax_percent": payload.tax_percent,
        "tax_amount": tax, "total": total, "status": payload.status, "ship_type": payload.ship_type,
        "currency": payload.currency, "notes": payload.notes, "created_at": now_iso(),
    }
    res = sb.table("invoices").insert(header).execute()
    invoice = res.data[0]

    line_rows = [{
        "invoice_id": invoice["id"], "product_id": it.product_id,
        "part_number": it.sku, "description": it.name, "qty": it.qty, "unit_price": it.unit_price,
    } for it in payload.items]
    if line_rows:
        sb.table("invoice_lines").insert(line_rows).execute()

    for it in payload.items:
        if it.product_id:
            await _bump_inventory(sb, it.product_id, -it.qty, invoice["id"])

    lines = sb.table("invoice_lines").select("*").eq("invoice_id", invoice["id"]).execute().data or []
    return _to_legacy(invoice, lines)


@router.put("/{iid}")
async def update_invoice(iid: str, payload: InvoiceCreate, staff=Depends(require_staff_or_admin)):
    sb = get_service_client()
    existing = sb.table("invoices").select("*").eq("id", iid).execute().data
    if not existing:
        raise HTTPException(status_code=404, detail="Invoice not found")
    existing = existing[0]
    dup = sb.table("invoices").select("id").eq("invoice_number", payload.invoice_number).neq("id", iid).execute()
    if dup.data:
        raise HTTPException(status_code=400, detail="Invoice number already exists")

    old_lines = sb.table("invoice_lines").select("*").eq("invoice_id", iid).execute().data or []
    old_qty = {}
    for li in old_lines:
        if li.get("product_id"):
            old_qty[li["product_id"]] = old_qty.get(li["product_id"], 0) + (li.get("qty") or 0)
    new_qty = {}
    for it in payload.items:
        if it.product_id:
            new_qty[it.product_id] = new_qty.get(it.product_id, 0) + it.qty

    for pid in set(list(old_qty) + list(new_qty)):
        net = old_qty.get(pid, 0) - new_qty.get(pid, 0)  # positive => stock returns
        if net < 0:
            avail = await _available_qty(sb, pid)
            if avail < -net:
                raise HTTPException(status_code=400, detail=f"Not enough stock for product {pid}: {avail} available")

    subtotal, tax, total = _totals(payload.items, payload.tax_percent)
    customer_id = await find_party_id_by_name(sb, "customers", payload.customer)
    sb.table("invoices").update({
        "customer": payload.customer, "customer_id": customer_id, "subtotal": subtotal,
        "tax_percent": payload.tax_percent, "tax_amount": tax, "total": total,
        "status": payload.status, "ship_type": payload.ship_type, "currency": payload.currency,
        "notes": payload.notes,
    }).eq("id", iid).execute()

    sb.table("invoice_lines").delete().eq("invoice_id", iid).execute()
    line_rows = [{
        "invoice_id": iid, "product_id": it.product_id,
        "part_number": it.sku, "description": it.name, "qty": it.qty, "unit_price": it.unit_price,
    } for it in payload.items]
    if line_rows:
        sb.table("invoice_lines").insert(line_rows).execute()

    for pid in set(list(old_qty) + list(new_qty)):
        net = old_qty.get(pid, 0) - new_qty.get(pid, 0)
        if net != 0:
            await _bump_inventory(sb, pid, net, iid)

    invoice = sb.table("invoices").select("*").eq("id", iid).execute().data[0]
    lines = sb.table("invoice_lines").select("*").eq("invoice_id", iid).execute().data or []
    return _to_legacy(invoice, lines)


@router.delete("/{iid}")
async def delete_invoice(iid: str, staff=Depends(require_staff_or_admin)):
    sb = get_service_client()
    lines = sb.table("invoice_lines").select("*").eq("invoice_id", iid).execute().data or []
    for li in lines:
        if li.get("product_id"):
            await _bump_inventory(sb, li["product_id"], li.get("qty", 0) or 0, iid)
    sb.table("invoice_lines").delete().eq("invoice_id", iid).execute()
    sb.table("invoices").delete().eq("id", iid).execute()
    return {"ok": True}


class InvoiceStatusReq(BaseModel):
    status: str


@router.patch("/{iid}/status")
async def set_invoice_status(iid: str, payload: InvoiceStatusReq, staff=Depends(require_staff_or_admin)):
    sb = get_service_client()
    sb.table("invoices").update({"status": payload.status}).eq("id", iid).execute()
    invoice = sb.table("invoices").select("*").eq("id", iid).execute().data[0]
    lines = sb.table("invoice_lines").select("*").eq("invoice_id", iid).execute().data or []
    return _to_legacy(invoice, lines)


@router.get("/{iid}/excel")
async def invoice_excel(iid: str, staff=Depends(require_staff_or_admin)):
    sb = get_service_client()
    invoice = sb.table("invoices").select("*").eq("id", iid).execute().data
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    invoice = invoice[0]
    lines = sb.table("invoice_lines").select("*").eq("invoice_id", iid).execute().data or []

    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice"
    ws["A1"] = "TAX INVOICE / فاتورة ضريبية"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Invoice No / رقم الفاتورة: {invoice.get('invoice_number','')}"
    ws["A3"] = f"Customer / العميل: {invoice.get('customer','')}"
    ws["A4"] = f"Date / التاريخ: {(invoice.get('invoice_date') or '')[:10]}"
    headers = [
        "S.NO رقم", "Part No. رقم القطع", "Description التفاصيل", "QTY/Pcs. العدد",
        "AED/Pc. سعر الوحدة", "Taxable Amount AED المبلغ", "VAT Rate % معدل الضريبة",
        "VAT Amount AED الضريبة", "TOTAL (incl VAT) AED المبلغ الأجمالي",
    ]
    hrow = 6
    fill = PatternFill(start_color="2C302B", end_color="2C302B", fill_type="solid")
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=hrow, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF", size=9)
        cell.fill = fill
        cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
    rate = invoice.get("tax_percent", 0) or 0
    r = hrow + 1
    for i, li in enumerate(lines, start=1):
        taxable = (li.get("qty", 0) or 0) * (li.get("unit_price", 0) or 0)
        vat = taxable * rate / 100.0
        ws.cell(row=r, column=1, value=i)
        ws.cell(row=r, column=2, value=li.get("part_number", ""))
        ws.cell(row=r, column=3, value=li.get("description", ""))
        ws.cell(row=r, column=4, value=li.get("qty", 0))
        ws.cell(row=r, column=5, value=round(li.get("unit_price", 0), 2))
        ws.cell(row=r, column=6, value=round(taxable, 2))
        ws.cell(row=r, column=7, value=rate)
        ws.cell(row=r, column=8, value=round(vat, 2))
        ws.cell(row=r, column=9, value=round(taxable + vat, 2))
        r += 1
    ws.cell(row=r + 1, column=8, value="Subtotal").font = Font(bold=True)
    ws.cell(row=r + 1, column=9, value=round(invoice.get("subtotal", 0), 2))
    ws.cell(row=r + 2, column=8, value="VAT").font = Font(bold=True)
    ws.cell(row=r + 2, column=9, value=round(invoice.get("tax_amount", 0), 2))
    ws.cell(row=r + 3, column=8, value="TOTAL").font = Font(bold=True)
    ws.cell(row=r + 3, column=9, value=round(invoice.get("total", 0), 2))
    widths = [8, 16, 32, 10, 14, 18, 14, 16, 20]
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=hrow, column=c).column_letter].width = w
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=invoice-{invoice.get('invoice_number','')}.xlsx"},
    )
