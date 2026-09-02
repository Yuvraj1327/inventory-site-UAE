import csv
import io
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from fastapi.responses import Response
from pydantic import BaseModel
from typing import List, Optional
from openpyxl import load_workbook

from app.core.security import require_staff_or_admin
from app.core.supabase_client import get_service_client
from app.services.compat import with_legacy_id, clean_list, now_iso, find_party_id_by_name
from app.services.audit import log_action
from app.services.ai_provider import get_extraction_provider
from app.services.file_storage import save_purchase_invoice, read_purchase_invoice

router = APIRouter(prefix="/api/purchases", tags=["purchases"])


class PurchaseItem(BaseModel):
    name: str = ""
    sku: str = ""
    qty: float = 0.0
    unit_cost: float = 0.0


class PurchaseCreate(BaseModel):
    supplier: str = ""
    ref: str = ""
    date: Optional[str] = None
    items: List[PurchaseItem] = []
    notes: str = ""


def _to_legacy(purchase: dict, lines: list[dict]) -> dict:
    row = with_legacy_id(purchase)
    row["date"] = purchase.get("purchase_date")
    row["ref"] = purchase.get("purchase_ref")
    row["items"] = [
        {"id": li.get("id"), "name": li.get("description", ""), "sku": li.get("part_number", ""),
         "qty": li.get("qty", 0), "unit_cost": li.get("unit_cost", 0),
         "sold_disposition": li.get("sold_disposition"), "disposition_decided_at": li.get("disposition_decided_at")}
        for li in lines
    ]
    return row


async def _adjust_stock(sb, name: str, sku: str, qty: float):
    """Mirrors the old Mongo `_adjust_stock`: find product by sku (else name), bump inventory, create if missing."""
    prod = None
    if sku:
        r = sb.table("products").select("id").eq("part_number", sku).limit(1).execute()
        prod = r.data[0] if r.data else None
    if not prod and name:
        r = sb.table("products").select("id").eq("description", name).limit(1).execute()
        prod = r.data[0] if r.data else None

    if prod:
        pid = prod["id"]
        inv = sb.table("inventory").select("*").eq("product_id", pid).execute().data
        if inv:
            new_qty = (inv[0].get("available_qty", 0) or 0) + qty
            sb.table("inventory").update({"available_qty": max(new_qty, 0), "updated_at": now_iso()}).eq("product_id", pid).execute()
        else:
            sb.table("inventory").insert({"product_id": pid, "available_qty": max(qty, 0)}).execute()
    else:
        ins = sb.table("products").insert({
            "description": name or sku, "part_number": sku, "unit_cost": 0.0,
            "default_selling_price": 0.0, "low_stock_threshold": 5.0, "created_at": now_iso(),
        }).execute()
        pid = ins.data[0]["id"]
        sb.table("inventory").insert({"product_id": pid, "available_qty": max(qty, 0)}).execute()
    return pid


@router.get("")
async def list_purchases(supplier: Optional[str] = None, staff=Depends(require_staff_or_admin)):
    sb = get_service_client()
    q = sb.table("purchases").select("*")
    if supplier:
        q = q.eq("supplier", supplier)
    purchases = q.order("purchase_date", desc=True).execute().data or []
    if not purchases:
        return []
    ids = [p["id"] for p in purchases]
    lines = sb.table("purchase_lines").select("*").in_("purchase_id", ids).execute().data or []
    lines_by_pid = {}
    for li in lines:
        lines_by_pid.setdefault(li["purchase_id"], []).append(li)
    return [_to_legacy(p, lines_by_pid.get(p["id"], [])) for p in purchases]


@router.post("")
async def create_purchase(payload: PurchaseCreate, staff=Depends(require_staff_or_admin)):
    sb = get_service_client()
    supplier_id = await find_party_id_by_name(sb, "suppliers", payload.supplier)
    total = sum(i.qty * i.unit_cost for i in payload.items)
    header = {
        "supplier": payload.supplier, "supplier_id": supplier_id,
        "purchase_ref": payload.ref, "purchase_date": (payload.date or now_iso())[:10],
        "total": total, "notes": payload.notes, "status": "received",
        "created_at": now_iso(),
    }
    res = sb.table("purchases").insert(header).execute()
    purchase = res.data[0]

    line_rows = []
    for it in payload.items:
        pid = await _adjust_stock(sb, it.name, it.sku, it.qty)
        if it.unit_cost:
            sb.table("products").update({"unit_cost": it.unit_cost}).eq("id", pid).execute()
        line_rows.append({
            "purchase_id": purchase["id"], "product_id": pid,
            "part_number": it.sku, "description": it.name,
            "qty": it.qty, "unit_cost": it.unit_cost,
        })
        sb.table("inventory_movements").insert({
            "product_id": pid, "movement_type": "purchase_in", "quantity": it.qty,
            "unit_cost": it.unit_cost, "reference_type": "purchase", "reference_id": purchase["id"],
        }).execute()
    if line_rows:
        sb.table("purchase_lines").insert(line_rows).execute()

    lines = sb.table("purchase_lines").select("*").eq("purchase_id", purchase["id"]).execute().data or []
    return _to_legacy(purchase, lines)


@router.delete("/{pid}")
async def delete_purchase(pid: str, staff=Depends(require_staff_or_admin)):
    sb = get_service_client()
    lines = sb.table("purchase_lines").select("*").eq("purchase_id", pid).execute().data or []
    for li in lines:
        if li.get("product_id"):
            await _adjust_stock(sb, li.get("description", ""), li.get("part_number", ""), -(li.get("qty", 0) or 0))
    sb.table("purchase_lines").delete().eq("purchase_id", pid).execute()
    sb.table("purchases").delete().eq("id", pid).execute()
    return {"ok": True}


@router.post("/upload")
async def upload_purchases(file: UploadFile = File(...), supplier: str = Form(""), staff=Depends(require_staff_or_admin)):
    content = await file.read()
    fname = (file.filename or "").lower()
    rows = []
    if fname.endswith(".csv"):
        reader = csv.DictReader(io.StringIO(content.decode("utf-8", errors="ignore")))
        for r in reader:
            rows.append({(k or "").strip().lower(): v for k, v in r.items()})
    else:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        data = list(ws.iter_rows(values_only=True))
        headers = [str(h).strip().lower() if h else "" for h in data[0]]
        for r in data[1:]:
            if not any(r):
                continue
            rows.append({headers[i]: r[i] for i in range(len(headers)) if headers[i]})

    def g(row, *keys):
        for k in keys:
            if k in row and row[k] not in (None, ""):
                return row[k]
        return ""

    def fnum(v):
        try:
            return float(str(v).replace(",", "").strip() or 0)
        except Exception:
            return 0.0

    items = []
    for r in rows:
        name = str(g(r, "description", "name", "product", "item") or "").strip()
        sku = str(g(r, "part number", "part no", "part no.", "sku", "code") or "").strip()
        if not name and not sku:
            continue
        items.append(PurchaseItem(
            name=name or sku, sku=sku,
            qty=fnum(g(r, "confirmed qty", "confirmed quantity", "qty", "quantity")),
            unit_cost=fnum(g(r, "price", "unit_cost", "cost")),
        ))
    if not items:
        raise HTTPException(status_code=400, detail="No valid rows found. Expected columns: Part Number, Description, Confirmed Qty, Price, Total")
    result = await create_purchase(
        PurchaseCreate(supplier=supplier, ref=f"UPLOAD-{file.filename}", items=items), staff
    )
    return {"created": True, "items": len(items), "purchase": result}


class ReorderItem(BaseModel):
    product_id: str
    qty: float = 0.0


class ReorderReq(BaseModel):
    supplier: str = ""
    items: List[ReorderItem] = []


@router.post("/reorder")
async def reorder(payload: ReorderReq, staff=Depends(require_staff_or_admin)):
    sb = get_service_client()
    items = []
    for r in payload.items:
        if r.qty <= 0:
            continue
        prod = sb.table("products").select("*").eq("id", r.product_id).execute().data
        if not prod:
            continue
        p = prod[0]
        items.append(PurchaseItem(name=p.get("description", ""), sku=p.get("part_number", ""), qty=r.qty, unit_cost=p.get("unit_cost", 0)))
    if not items:
        raise HTTPException(status_code=400, detail="No items to reorder")
    return await create_purchase(PurchaseCreate(supplier=payload.supplier, ref="REORDER", items=items), staff)


@router.post("/scan")
async def scan_purchase(file: UploadFile = File(...), staff=Depends(require_staff_or_admin)):
    # AI invoice extraction is implemented in Phase 4. The old Emergent LLM
    # integration was removed in Phase 1 and not yet replaced.
    raise HTTPException(status_code=503, detail="AI purchase-invoice scanning lands in Phase 4.")


# =====================================================================
# PHASE 4 — Purchase Confirmation workflow
#
# The endpoints above (POST /purchases, /reorder, /upload) remain as a
# lightweight "quick entry" path that behaves like disposition='none'
# (goods go straight to available stock) — unchanged, still used by the
# existing Purchases.jsx quick-add dialog.
#
# The endpoints below implement the fuller spec'd workflow: a registered
# supplier + invoice reference is required up front, an invoice document
# can be attached and (optionally) AI-extracted, and receiving is a
# separate, explicit No/Full/Partial decision per line (Phase 5's
# "goods already sold" behavior) rather than an automatic stock bump.
# =====================================================================


class ConfirmItem(BaseModel):
    part_number: str
    description: str = ""
    qty: float
    unit_cost: float


class PurchaseConfirmReq(BaseModel):
    supplier_id: str
    supplier_invoice_number: str
    purchase_date: Optional[str] = None
    items: List[ConfirmItem] = []
    notes: str = ""


@router.post("/confirm")
async def confirm_purchase(payload: PurchaseConfirmReq, staff=Depends(require_staff_or_admin)):
    sb = get_service_client()

    supplier = sb.table("suppliers").select("id, name").eq("id", payload.supplier_id).execute().data
    if not supplier:
        raise HTTPException(status_code=400, detail="Supplier must be a registered supplier. Register them under Suppliers first.")
    if not payload.supplier_invoice_number.strip():
        raise HTTPException(status_code=400, detail="Supplier invoice number is required")
    if not payload.items:
        raise HTTPException(status_code=400, detail="Add at least one purchase line")
    for it in payload.items:
        if it.qty <= 0:
            raise HTTPException(status_code=400, detail=f"Quantity must be > 0 for {it.part_number}")
        if it.unit_cost < 0:
            raise HTTPException(status_code=400, detail=f"Unit cost cannot be negative for {it.part_number}")

    dup = sb.table("purchases").select("id").eq("supplier_id", payload.supplier_id).eq("supplier_invoice_number", payload.supplier_invoice_number.strip()).execute()
    if dup.data:
        raise HTTPException(status_code=400, detail="This supplier invoice number has already been recorded")

    total = sum(i.qty * i.unit_cost for i in payload.items)
    header = {
        "supplier": supplier[0]["name"], "supplier_id": payload.supplier_id,
        "supplier_invoice_number": payload.supplier_invoice_number.strip(),
        "purchase_date": (payload.purchase_date or now_iso())[:10],
        "total": total, "notes": payload.notes, "status": "pending",
        "extraction_status": "not_attempted", "created_at": now_iso(),
    }
    res = sb.table("purchases").insert(header).execute()
    purchase = res.data[0]

    line_rows = []
    for it in payload.items:
        existing_product = sb.table("products").select("id").eq("part_number", it.part_number).limit(1).execute().data
        if existing_product:
            pid = existing_product[0]["id"]
            sb.table("products").update({"unit_cost": it.unit_cost}).eq("id", pid).execute()
        else:
            # Controlled product creation: server-validated fields only, never
            # arbitrary client-supplied columns.
            ins = sb.table("products").insert({
                "part_number": it.part_number, "description": it.description or it.part_number,
                "unit_cost": it.unit_cost, "default_selling_price": 0.0,
                "low_stock_threshold": 5.0, "created_at": now_iso(),
            }).execute()
            pid = ins.data[0]["id"]
            sb.table("inventory").insert({"product_id": pid, "available_qty": 0}).execute()

        line_rows.append({
            "purchase_id": purchase["id"], "product_id": pid, "part_number": it.part_number,
            "description": it.description, "qty": it.qty, "unit_cost": it.unit_cost,
            "sold_disposition": "none",
        })
    inserted_lines = sb.table("purchase_lines").insert(line_rows).execute().data or []

    log_action(sb, staff.get("id"), "purchase.confirm", "purchase", purchase["id"], {
        "supplier": supplier[0]["name"], "invoice_number": payload.supplier_invoice_number, "line_count": len(inserted_lines),
    })
    return with_legacy_id({**purchase, "items": inserted_lines})


@router.post("/{pid}/upload-invoice")
async def upload_invoice(pid: str, file: UploadFile = File(...), staff=Depends(require_staff_or_admin)):
    sb = get_service_client()
    existing = sb.table("purchases").select("id").eq("id", pid).execute().data
    if not existing:
        raise HTTPException(status_code=404, detail="Purchase not found")
    path, original_name = await save_purchase_invoice(file)
    sb.table("purchases").update({"invoice_file_path": path}).eq("id", pid).execute()
    log_action(sb, staff.get("id"), "purchase.upload_invoice", "purchase", pid, {"filename": original_name})
    return {"ok": True, "filename": original_name}


@router.get("/{pid}/download-invoice")
async def download_invoice(pid: str, staff=Depends(require_staff_or_admin)):
    sb = get_service_client()
    row = sb.table("purchases").select("invoice_file_path").eq("id", pid).execute().data
    if not row or not row[0].get("invoice_file_path"):
        raise HTTPException(status_code=404, detail="No invoice file on this purchase")
    content = read_purchase_invoice(row[0]["invoice_file_path"])
    return Response(content=content, media_type="application/octet-stream")


@router.post("/{pid}/extract")
async def extract_invoice(pid: str, staff=Depends(require_staff_or_admin)):
    sb = get_service_client()
    row = sb.table("purchases").select("*").eq("id", pid).execute().data
    if not row or not row[0].get("invoice_file_path"):
        raise HTTPException(status_code=400, detail="Upload an invoice file first")
    content = read_purchase_invoice(row[0]["invoice_file_path"])

    provider = get_extraction_provider()
    result = await provider.extract(content, row[0].get("invoice_file_path", ""))
    sb.table("purchases").update({"extraction_status": result_status(result)}).eq("id", pid).execute()
    log_action(sb, staff.get("id"), "purchase.extract_attempt", "purchase", pid, {"ok": result.ok})
    return result.to_dict()


def result_status(result) -> str:
    if not result.ok:
        return "not_configured" if "not configured" in result.message.lower() else "needs_review"
    return "needs_review"


class ReceiveLine(BaseModel):
    purchase_line_id: str
    disposition: str  # "none" | "full" | "partial"
    order_line_id: Optional[str] = None
    qty: Optional[float] = None  # required for "partial"; ignored for "full" (uses full line qty) and "none"


class ReceiveReq(BaseModel):
    lines: List[ReceiveLine]


@router.post("/{pid}/receive")
async def receive_purchase(pid: str, payload: ReceiveReq, staff=Depends(require_staff_or_admin)):
    """
    Phase 5 "goods already sold after receipt" workflow. For each
    purchase line, one of:
      - none:    full received qty goes into available inventory.
      - full:    full received qty is allocated directly to the given
                 order line (never touches available stock).
      - partial: `qty` is allocated to the order line; the remainder
                 goes into available inventory.
    All quantity transitions are validated so nothing can go negative
    or exceed what was actually received.
    """
    sb = get_service_client()
    purchase = sb.table("purchases").select("*").eq("id", pid).execute().data
    if not purchase:
        raise HTTPException(status_code=404, detail="Purchase not found")

    purchase_lines = {li["id"]: li for li in (sb.table("purchase_lines").select("*").eq("purchase_id", pid).execute().data or [])}
    if not purchase_lines:
        raise HTTPException(status_code=400, detail="This purchase has no lines")

    for rl in payload.lines:
        li = purchase_lines.get(rl.purchase_line_id)
        if not li:
            raise HTTPException(status_code=400, detail=f"Purchase line {rl.purchase_line_id} not found on this purchase")
        if li.get("disposition_decided_at"):
            raise HTTPException(status_code=400, detail=f"Line {li.get('part_number')} has already been received/decided")
        if rl.disposition not in ("none", "full", "partial"):
            raise HTTPException(status_code=400, detail="disposition must be 'none', 'full', or 'partial'")
        if rl.disposition in ("full", "partial") and not rl.order_line_id:
            raise HTTPException(status_code=400, detail="order_line_id is required for full/partial allocation")
        if rl.disposition == "partial":
            if rl.qty is None or rl.qty <= 0:
                raise HTTPException(status_code=400, detail="qty is required and must be > 0 for a partial disposition")
            if rl.qty > li["qty"]:
                raise HTTPException(status_code=400, detail=f"Cannot allocate more than the received qty ({li['qty']}) for {li.get('part_number')}")

    for rl in payload.lines:
        li = purchase_lines[rl.purchase_line_id]
        pid_product = li.get("product_id")
        received_qty = li["qty"]
        allocate_qty = received_qty if rl.disposition == "full" else (rl.qty if rl.disposition == "partial" else 0)
        available_qty = received_qty - allocate_qty  # goes to stock

        if pid_product:
            sb.table("inventory_movements").insert({
                "product_id": pid_product, "movement_type": "purchase_in", "quantity": received_qty,
                "unit_cost": li.get("unit_cost"), "reference_type": "purchase", "reference_id": pid,
            }).execute()
            inv = sb.table("inventory").select("*").eq("product_id", pid_product).execute().data
            current = inv[0] if inv else None
            new_available = (current.get("available_qty", 0) if current else 0) + available_qty
            new_sold = (current.get("sold_qty", 0) if current else 0) + allocate_qty
            if current:
                sb.table("inventory").update({"available_qty": new_available, "sold_qty": new_sold, "updated_at": now_iso()}).eq("product_id", pid_product).execute()
            else:
                sb.table("inventory").insert({"product_id": pid_product, "available_qty": new_available, "sold_qty": new_sold}).execute()

            if allocate_qty:
                sb.table("inventory_movements").insert({
                    "product_id": pid_product, "movement_type": "allocation", "quantity": -allocate_qty,
                    "unit_cost": li.get("unit_cost"), "reference_type": "order", "reference_id": rl.order_line_id,
                }).execute()

        if allocate_qty and rl.order_line_id:
            ol = sb.table("order_lines").select("*").eq("id", rl.order_line_id).execute().data
            if ol:
                new_shipped = (ol[0].get("shipped_qty") or 0) + allocate_qty
                new_confirm = max(ol[0].get("confirm_qty") or 0, new_shipped)
                sb.table("order_lines").update({
                    "shipped_qty": new_shipped, "confirm_qty": new_confirm,
                    "status": "Shipped" if new_shipped >= new_confirm else "Confirmed",
                }).eq("id", rl.order_line_id).execute()

        sb.table("purchase_lines").update({
            "sold_disposition": rl.disposition, "allocated_order_line_id": rl.order_line_id,
            "allocated_qty": allocate_qty, "disposition_decided_at": now_iso(),
        }).eq("id", rl.purchase_line_id).execute()

        log_action(sb, staff.get("id"), "purchase_line.receive", "purchase_line", rl.purchase_line_id, {
            "disposition": rl.disposition, "allocated_qty": allocate_qty, "available_qty": available_qty,
        })

    remaining_undecided = sb.table("purchase_lines").select("id").eq("purchase_id", pid).is_("disposition_decided_at", "null").execute().data or []
    new_status = "partial" if remaining_undecided else "received"
    sb.table("purchases").update({"status": new_status}).eq("id", pid).execute()

    lines = sb.table("purchase_lines").select("*").eq("purchase_id", pid).execute().data or []
    return _to_legacy(sb.table("purchases").select("*").eq("id", pid).execute().data[0], lines)
