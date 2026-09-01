import csv
import io
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from pydantic import BaseModel
from openpyxl import load_workbook

from app.core.security import require_staff_or_admin
from app.core.supabase_client import get_service_client
from app.services.compat import now_iso

router = APIRouter(prefix="/api/products", tags=["products"])


class ProductCreate(BaseModel):
    name: str
    sku: str = ""
    stock: float = 0.0
    unit_cost: float = 0.0
    unit_price: float = 0.0
    low_stock_threshold: float = 5.0
    pc: str = ""
    superseded: str = ""
    superseded_from: str = ""
    model: str = ""
    availability: str = ""


def _to_legacy(product: dict, inv: dict | None) -> dict:
    return {
        "_id": product["id"], "id": product["id"],
        "name": product.get("description", ""), "sku": product.get("part_number", ""),
        "stock": (inv or {}).get("available_qty", 0),
        "unit_cost": product.get("unit_cost", 0), "unit_price": product.get("default_selling_price", 0),
        "low_stock_threshold": product.get("low_stock_threshold", 5),
        "pc": product.get("pc", ""), "superseded": product.get("superseded_reference", ""),
        "superseded_from": product.get("superseded_from", ""), "model": product.get("model", ""),
        "availability": product.get("availability", ""),
        "created_at": product.get("created_at"),
    }


def _from_legacy(payload: ProductCreate) -> dict:
    return {
        "description": payload.name, "part_number": payload.sku,
        "unit_cost": payload.unit_cost, "default_selling_price": payload.unit_price,
        "low_stock_threshold": payload.low_stock_threshold, "pc": payload.pc,
        "superseded_reference": payload.superseded, "superseded_from": payload.superseded_from,
        "model": payload.model, "availability": payload.availability,
    }


async def _upsert_inventory(sb, product_id: str, stock: float):
    existing = sb.table("inventory").select("product_id").eq("product_id", product_id).execute()
    if existing.data:
        sb.table("inventory").update({"available_qty": stock, "updated_at": now_iso()}).eq("product_id", product_id).execute()
    else:
        sb.table("inventory").insert({"product_id": product_id, "available_qty": stock}).execute()


@router.get("")
async def list_products(staff=Depends(require_staff_or_admin)):
    sb = get_service_client()
    products = sb.table("products").select("*").order("description").execute().data or []
    inv_rows = sb.table("inventory").select("*").execute().data or []
    inv_by_pid = {r["product_id"]: r for r in inv_rows}
    return [_to_legacy(p, inv_by_pid.get(p["id"])) for p in products]


@router.post("")
async def create_product(payload: ProductCreate, staff=Depends(require_staff_or_admin)):
    sb = get_service_client()
    data = _from_legacy(payload)
    data["created_at"] = now_iso()
    res = sb.table("products").insert(data).execute()
    row = res.data[0]
    await _upsert_inventory(sb, row["id"], payload.stock)
    inv = sb.table("inventory").select("*").eq("product_id", row["id"]).execute().data[0]
    return _to_legacy(row, inv)


@router.put("/{pid}")
async def update_product(pid: str, payload: ProductCreate, staff=Depends(require_staff_or_admin)):
    sb = get_service_client()
    data = _from_legacy(payload)
    sb.table("products").update(data).eq("id", pid).execute()
    await _upsert_inventory(sb, pid, payload.stock)
    row = sb.table("products").select("*").eq("id", pid).execute().data[0]
    inv = sb.table("inventory").select("*").eq("product_id", pid).execute().data[0]
    return _to_legacy(row, inv)


@router.get("/low-stock")
async def low_stock(staff=Depends(require_staff_or_admin)):
    sb = get_service_client()
    products = sb.table("products").select("*").execute().data or []
    inv_rows = sb.table("inventory").select("*").execute().data or []
    inv_by_pid = {r["product_id"]: r for r in inv_rows}
    out = []
    for p in products:
        inv = inv_by_pid.get(p["id"], {})
        if (inv.get("available_qty", 0) or 0) <= (p.get("low_stock_threshold", 0) or 0):
            out.append(_to_legacy(p, inv))
    return out


@router.delete("/{pid}")
async def delete_product(pid: str, staff=Depends(require_staff_or_admin)):
    sb = get_service_client()
    sb.table("inventory").delete().eq("product_id", pid).execute()
    sb.table("products").delete().eq("id", pid).execute()
    return {"ok": True}


@router.post("/upload")
async def upload_products(file: UploadFile = File(...), staff=Depends(require_staff_or_admin)):
    sb = get_service_client()
    content = await file.read()
    fname = (file.filename or "").lower()
    rows = []
    if fname.endswith(".csv"):
        reader = csv.DictReader(io.StringIO(content.decode("utf-8", errors="ignore")))
        for r in reader:
            rows.append({(k or "").strip().lower(): v for k, v in r.items()})
    else:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        data = list(ws.iter_rows(values_only=True))
        headers = [str(h).strip().lower() if h else "" for h in data[0]]
        for r in data[1:]:
            if not any(r):
                continue
            rows.append({headers[i]: r[i] for i in range(len(headers)) if headers[i]})

    def g(row, *keys):
        for k in keys:
            if k in row and row[k] not in (None, ""):
                return row[k]
        return ""

    def fnum(v):
        try:
            return float(str(v).replace(",", "").strip() or 0)
        except Exception:
            return 0.0

    count = 0
    for r in rows:
        sku = str(g(r, "part number", "part no", "part no.", "sku", "code") or "").strip()
        name = str(g(r, "description", "name", "item") or sku).strip()
        if not sku and not name:
            continue
        stock = fnum(g(r, "req qty", "req. qty", "qty", "quantity", "stock"))
        data = {
            "description": name, "part_number": sku,
            "unit_cost": fnum(g(r, "req. price", "req price", "cost", "unit_cost")),
            "default_selling_price": fnum(g(r, "net price", "price", "unit_price")),
            "pc": str(g(r, "pc") or ""),
            "superseded_reference": str(g(r, "superseded") or ""),
            "superseded_from": str(g(r, "superseded from") or ""),
            "model": str(g(r, "model") or ""),
            "availability": str(g(r, "availability") or ""),
        }
        existing = (
            sb.table("products").select("id").eq("part_number", sku).limit(1).execute()
            if sku else sb.table("products").select("id").eq("description", name).limit(1).execute()
        )
        if existing.data:
            pid = existing.data[0]["id"]
            sb.table("products").update(data).eq("id", pid).execute()
        else:
            data["low_stock_threshold"] = 5.0
            data["created_at"] = now_iso()
            res = sb.table("products").insert(data).execute()
            pid = res.data[0]["id"]
        await _upsert_inventory(sb, pid, stock)
        count += 1

    if count == 0:
        raise HTTPException(status_code=400, detail="No valid rows. Expected columns like: Part Number, Description, Req Qty, Req. price, Net Price")
    return {"imported": count}
